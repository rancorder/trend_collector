"""
trend_collector/collect.py
==========================
AI・エンジニアリング全般のトレンド収集スクリプト

収集ソース:
  - GitHub Trending（全言語・週次）
  - Qiita API（AI/ML/インフラ/セキュリティ/フロント/組み込み）
  - Zenn RSS（フィルタなし全件）
  - npm（AI/インフラ/ツール系パッケージ）

カテゴリ分類:
  AI/LLM / 組み込み・エッジ / インフラ・クラウド / フロントエンド /
  セキュリティ / DevOps / データ / その他

出力:
  - trend_output/report_YYYYMMDD.html
  - trend_output/report_YYYYMMDD.xlsx

実行:
  python trend_collector/collect.py
"""
from __future__ import annotations

import logging
import re
import time
from dataclasses import dataclass, field, asdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import requests
from bs4 import BeautifulSoup

log = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S")

OUTPUT_DIR = Path(__file__).parent.parent / "trend_output"
OUTPUT_DIR.mkdir(exist_ok=True)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/122.0.0.0 Safari/537.36",
    "Accept-Language": "ja,en-US;q=0.9",
}

# ─── カテゴリ定義 ──────────────────────────────────────────────────────
# タイトル・タグ・説明文に含まれるキーワードでカテゴリ分類
CATEGORY_RULES: list[tuple[str, list[str]]] = [
    ("🤖 AI / LLM",         ["llm","gpt","claude","gemini","生成ai","generative","agent","rag","fine-tun","transformer","diffusion","stable diffusion","whisper","tts","音声合成","mcp","langchain","llamaindex","ollama","vllm","hugging"]),
    ("📦 組み込み・エッジ",  ["組み込み","embedded","fpga","rtos","arm","マイコン","ファームウェア","firmware","iot","m2m","edgeai","edge ai","esp32","raspberry","arduino","zephyr","ros","lidar","センサ","sensor","カメラモジュール","npu","tinyml","coral","jetson","hailo"]),
    ("☁️ インフラ・クラウド",["kubernetes","k8s","docker","terraform","aws","gcp","azure","infra","infrastructure","serverless","lambda","cloud","cdn","nginx","istio","helm","argocd","gitops"]),
    ("🎨 フロントエンド",    ["react","vue","next.js","nuxt","svelte","typescript","tailwind","css","ui/ux","frontend","フロント","wasm","webassembly","web components","shadcn"]),
    ("🔐 セキュリティ",      ["security","セキュリティ","脆弱性","vulnerability","pentest","zero-day","cve","sast","dast","siem","zero trust","oauth","認証","暗号","encryption"]),
    ("⚙️ DevOps / MLOps",   ["ci/cd","devops","mlops","github actions","gitlab","observability","monitoring","datadog","opentelemetry","prometheus","grafana","sre","platform eng"]),
    ("📊 データ・分析",      ["データ分析","data","spark","dbt","bigquery","snowflake","airflow","pandas","polars","duckdb","etl","warehouse","analytics","bi","tableau","sql"]),
    ("🦀 言語・ランタイム",  ["rust","go","zig","bun","deno","python","kotlin","swift","ruby","llvm","compiler","runtime","wasm"]),
]

def classify(text: str) -> str:
    t = text.lower()
    for cat, kws in CATEGORY_RULES:
        if any(k in t for k in kws):
            return cat
    return "💡 その他"


# ─── データモデル ──────────────────────────────────────────────────────
@dataclass
class TrendItem:
    source:      str
    title:       str
    url:         str
    category:    str  = ""
    description: str  = ""
    language:    str  = ""
    stars:       int  = 0
    likes:       int  = 0
    tags:        list = field(default_factory=list)
    downloads:   int  = 0
    is_notable:  bool = False  # 特に注目すべき項目

@dataclass
class TrendReport:
    generated_at: str
    github_items: list[TrendItem] = field(default_factory=list)
    qiita_items:  list[TrendItem] = field(default_factory=list)
    zenn_items:   list[TrendItem] = field(default_factory=list)
    npm_items:    list[TrendItem] = field(default_factory=list)
    new_keywords: list[str]       = field(default_factory=list)


# ─── HTTP取得 ─────────────────────────────────────────────────────────
def _get(url: str, params: dict = None, timeout: int = 15) -> Optional[requests.Response]:
    try:
        resp = requests.get(url, headers=HEADERS, params=params, timeout=timeout)
        resp.raise_for_status()
        return resp
    except Exception as e:
        log.warning(f"  取得失敗: {url[:80]} → {e}")
        return None


# ─── GitHub Trending ──────────────────────────────────────────────────
def collect_github_trending(since: str = "weekly") -> list[TrendItem]:
    log.info("  [GitHub Trending] 収集中...")
    items = []
    # 全言語 + Python + TypeScript + Rust + Go の5URL取得して合算
    urls = [
        f"https://github.com/trending?since={since}",
        f"https://github.com/trending/python?since={since}",
        f"https://github.com/trending/typescript?since={since}",
        f"https://github.com/trending/rust?since={since}",
        f"https://github.com/trending/go?since={since}",
    ]
    seen = set()
    for url in urls:
        resp = _get(url)
        if not resp:
            continue
        soup = BeautifulSoup(resp.text, "lxml")
        for repo in soup.select("article.Box-row"):
            name_tag = repo.select_one("h2.h3 a, h2.lh-condensed a, h2 a")
            if not name_tag:
                continue
            href = name_tag.get("href", "")
            if not href or href in seen:
                continue
            seen.add(href)
            title = href.strip("/").replace("/", " / ")
            desc_tag = repo.select_one("p")
            desc = desc_tag.get_text(strip=True) if desc_tag else ""
            lang_tag = repo.select_one("[itemprop='programmingLanguage']")
            lang = lang_tag.get_text(strip=True) if lang_tag else ""
            stars = 0
            stars_tag = repo.select_one("a[href$='/stargazers']")
            if stars_tag:
                raw = re.sub(r"[,\s]", "", stars_tag.get_text(separator=" ", strip=True))
                m = re.search(r"(\d+)", raw)
                if m:
                    try: stars = int(m.group(1))
                    except: pass
            cat = classify(title + " " + desc + " " + lang)
            item = TrendItem(
                source="GitHub", title=title,
                url=f"https://github.com{href}",
                category=cat, description=desc, language=lang, stars=stars,
            )
            items.append(item)
        time.sleep(0.5)

    # スター数降順
    items.sort(key=lambda x: x.stars, reverse=True)
    log.info(f"    → {len(items)} 件")
    return items


# ─── Qiita ────────────────────────────────────────────────────────────
def collect_qiita(per_page: int = 30) -> list[TrendItem]:
    log.info("  [Qiita] 収集中...")
    week_ago = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")

    # 幅広いカテゴリで収集
    queries = [
        # AI/LLM
        "tag:LLM OR tag:生成AI OR tag:ChatGPT OR tag:Claude OR tag:RAG OR tag:AIエージェント",
        "tag:機械学習 OR tag:深層学習 OR tag:PyTorch OR tag:TensorFlow",
        # インフラ・クラウド
        "tag:Kubernetes OR tag:Docker OR tag:AWS OR tag:Terraform OR tag:GitHubActions",
        # 組み込み・エッジ
        "tag:組み込み OR tag:IoT OR tag:FPGA OR tag:RTOS OR tag:ROS OR tag:EdgeAI",
        # フロント・言語
        "tag:TypeScript OR tag:React OR tag:Next.js OR tag:Rust OR tag:Go",
        # セキュリティ・DevOps
        "tag:セキュリティ OR tag:脆弱性 OR tag:MLOps OR tag:DevOps",
    ]
    items = []
    seen = set()
    for q in queries:
        resp = _get(
            "https://qiita.com/api/v2/items",
            params={"query": f"{q} created:>={week_ago}", "per_page": per_page, "page": 1},
        )
        if not resp:
            continue
        for art in resp.json():
            if art["id"] in seen:
                continue
            seen.add(art["id"])
            tags = [t["name"] for t in art.get("tags", [])]
            cat  = classify(art["title"] + " " + " ".join(tags))
            item = TrendItem(
                source="Qiita", title=art["title"],
                url=art["url"], category=cat, tags=tags,
                likes=art.get("likes_count", 0),
            )
            items.append(item)
        time.sleep(0.5)

    items.sort(key=lambda x: x.likes, reverse=True)
    log.info(f"    → {len(items)} 件（重複除去後）")
    return items[:50]


# ─── Zenn ─────────────────────────────────────────────────────────────
def collect_zenn() -> list[TrendItem]:
    log.info("  [Zenn] 収集中...")
    try:
        import feedparser
    except ImportError:
        log.warning("    feedparser未インストール → スキップ")
        return []

    # トレンドRSS + トピック別RSS
    feeds = [
        "https://zenn.dev/feed",                      # 全体トレンド
        "https://zenn.dev/topics/llm/feed",           # LLM
        "https://zenn.dev/topics/ai/feed",            # AI
        "https://zenn.dev/topics/typescript/feed",    # TypeScript
        "https://zenn.dev/topics/rust/feed",          # Rust
        "https://zenn.dev/topics/kubernetes/feed",    # Kubernetes
        "https://zenn.dev/topics/security/feed",      # セキュリティ
    ]
    items = []
    seen = set()
    for feed_url in feeds:
        feed = feedparser.parse(feed_url)
        for entry in feed.entries[:20]:
            url = entry.get("link", "")
            if url in seen:
                continue
            seen.add(url)
            title = entry.get("title", "")
            desc  = BeautifulSoup(entry.get("summary", ""), "lxml").get_text()[:150]
            tags  = [t.get("term","") for t in entry.get("tags", [])]
            cat   = classify(title + " " + " ".join(tags) + " " + feed_url)
            item  = TrendItem(
                source="Zenn", title=title,
                url=url, category=cat, description=desc, tags=tags,
            )
            items.append(item)
        time.sleep(0.3)

    log.info(f"    → {len(items)} 件")
    return items


# ─── npm ──────────────────────────────────────────────────────────────
def collect_npm_trends() -> list[TrendItem]:
    log.info("  [npm] 収集中...")
    # AI・インフラ・ツール系に拡大
    keywords = [
        "llm", "ai-sdk", "openai", "anthropic", "langchain",
        "edge-ai", "onnx", "tensorflow",
        "iot", "embedded",
        "kubernetes", "docker", "terraform",
        "vite", "esbuild", "bun",
    ]
    all_pkgs = []
    seen = set()
    for kw in keywords:
        resp = _get(
            "https://registry.npmjs.org/-/v1/search",
            params={"text": kw, "size": 8, "quality": 0.5, "popularity": 1.0},
        )
        if not resp:
            continue
        for obj in resp.json().get("objects", []):
            pkg = obj["package"]
            name = pkg["name"]
            if name not in seen:
                seen.add(name)
                all_pkgs.append(pkg)
        time.sleep(0.4)

    # DL数取得（スコープ付きは個別、通常は一括）
    dl_map: dict[str, int] = {}
    normal = [p["name"] for p in all_pkgs if not p["name"].startswith("@")]
    scoped = [p["name"] for p in all_pkgs if p["name"].startswith("@")]

    for i in range(0, len(normal), 10):
        chunk = normal[i:i+10]
        resp = _get("https://api.npmjs.org/downloads/point/last-week/" + ",".join(chunk))
        if resp:
            data = resp.json()
            if isinstance(data, dict) and "downloads" in data and len(chunk) == 1:
                dl_map[chunk[0]] = data.get("downloads", 0)
            elif isinstance(data, dict):
                for k, v in data.items():
                    if isinstance(v, dict):
                        dl_map[k] = v.get("downloads", 0)
        time.sleep(0.8)

    for name in scoped[:15]:
        encoded = name.replace("/", "%2F")
        resp = _get(f"https://api.npmjs.org/downloads/point/last-week/{encoded}")
        if resp:
            try: dl_map[name] = resp.json().get("downloads", 0)
            except: pass
        time.sleep(1.2)

    items = []
    for pkg in all_pkgs:
        name = pkg["name"]
        tags = pkg.get("keywords", [])[:5]
        cat  = classify(name + " " + pkg.get("description","") + " " + " ".join(tags))
        item = TrendItem(
            source="npm", title=name,
            url=f"https://www.npmjs.com/package/{name}",
            category=cat, description=pkg.get("description","")[:100],
            downloads=dl_map.get(name, 0), tags=tags,
        )
        items.append(item)

    items.sort(key=lambda x: x.downloads, reverse=True)
    log.info(f"    → {len(items)} 件")
    return items[:30]


# ─── 注目キーワード検出 ────────────────────────────────────────────────
# パターン → 正規化名称（タイトル・タグのみ対象）
WATCH_PATTERNS: dict[str, str] = {
    r"risc-?v":                        "RISC-V",
    r"rust.{0,10}embed":               "Rust組み込み",
    r"wasm.{0,10}edge":                "WASM Edge",
    r"tinyml":                         "TinyML",
    r"mlops":                          "MLOps",
    r"llm.{0,10}edge|edge.{0,10}llm":  "Edge LLM",
    r"npu":                            "NPU",
    r"ros2":                           "ROS2",
    r"autosar":                        "AUTOSAR",
    r"onnx.?runtime":                  "ONNX Runtime",
    r"tflite":                         "TFLite",
    r"hailo":                          "Hailo",
    r"jetson":                         "Jetson",
    r"esp32":                          "ESP32",
    r"mcp\b":                          "MCP (Model Context Protocol)",
    r"vibe.?cod":                       "Vibe Coding",
    r"cursor\b":                       "Cursor IDE",
    r"claude.?code":                   "Claude Code",
    r"opentelemetry":                  "OpenTelemetry",
    r"platform.?engineer":             "Platform Engineering",
    r"digital.?twin|デジタルツイン":    "デジタルツイン",
    r"agentic":                        "Agentic AI",
    r"reasoning.?model":               "Reasoning Model",
    r"multimodal":                     "Multimodal",
}

def _has_notable(text: str) -> bool:
    t = text.lower()
    return any(re.search(p, t) for p in WATCH_PATTERNS)

def _extract_notable_keywords(items: list[TrendItem]) -> list[str]:
    found = set()
    for item in items:
        text = (item.title + " " + " ".join(item.tags)).lower()
        for pattern, normalized in WATCH_PATTERNS.items():
            if re.search(pattern, text):
                found.add(normalized)
    return sorted(found)


# ─── HTML出力 ─────────────────────────────────────────────────────────
def _render_html(report: TrendReport) -> str:
    all_items = report.github_items + report.qiita_items + report.zenn_items + report.npm_items

    # カテゴリ別に集計
    from collections import defaultdict, Counter
    cat_counts = Counter(i.category for i in all_items)

    def _badge(source):
        colors = {"GitHub":"#24292e","Qiita":"#55C500","Zenn":"#3EA8FF","npm":"#CB3837"}
        c = colors.get(source, "#666")
        return f'<span style="background:{c};color:#fff;padding:1px 7px;border-radius:10px;font-size:11px;font-weight:bold;">{source}</span>'

    def _cat_badge(cat):
        return f'<span style="background:#EEF2FF;color:#3730A3;padding:1px 6px;border-radius:4px;font-size:10px;font-weight:bold;">{cat}</span>'

    def _item_rows(items):
        if not items:
            return "<p style='color:#999;font-size:12px;'>データなし</p>"
        rows = []
        for it in items[:25]:
            tags_html = " ".join(
                f'<span style="background:#F3F4F6;padding:1px 5px;border-radius:3px;font-size:10px;">{t}</span>'
                for t in (it.tags or [it.language])[:4] if t
            )
            metric = ""
            if it.stars:     metric = f'⭐ {it.stars:,}'
            elif it.likes:   metric = f'♡ {it.likes:,}'
            elif it.downloads: metric = f'↓ {it.downloads:,}/週'
            notable = ' <span style="background:#DC2626;color:#fff;padding:1px 5px;border-radius:8px;font-size:10px;">注目</span>' if it.is_notable else ""
            rows.append(f"""
            <tr>
              <td style="padding:8px 10px;border-bottom:1px solid #F3F4F6;">
                {_badge(it.source)} {_cat_badge(it.category)}{notable}<br>
                <a href="{it.url}" target="_blank"
                   style="color:#1D4ED8;font-weight:bold;font-size:13px;">{it.title}</a><br>
                <span style="color:#6B7280;font-size:11px;">{it.description[:90]}</span><br>
                <div style="margin-top:3px;">{tags_html}</div>
              </td>
              <td style="padding:8px 10px;border-bottom:1px solid #F3F4F6;text-align:right;
                         font-size:12px;color:#374151;white-space:nowrap;vertical-align:top;">
                {metric}
              </td>
            </tr>""")
        return f'<table style="width:100%;border-collapse:collapse;">{"".join(rows)}</table>'

    # カテゴリサマリー
    cat_summary = ""
    for cat, cnt in sorted(cat_counts.items(), key=lambda x: -x[1]):
        bar_w = min(int(cnt / max(cat_counts.values()) * 180), 180)
        cat_summary += f"""
        <div style="display:flex;align-items:center;gap:10px;margin:4px 0;">
          <div style="width:160px;font-size:12px;color:#374151;">{cat}</div>
          <div style="background:#DBEAFE;height:14px;width:{bar_w}px;border-radius:3px;"></div>
          <div style="font-size:12px;color:#6B7280;">{cnt}件</div>
        </div>"""

    # 注目キーワード
    notable_kws = _extract_notable_keywords(all_items)
    kw_chips = "".join(
        f'<span style="background:#FEF3C7;border:1px solid #F59E0B;color:#92400E;'
        f'padding:3px 10px;border-radius:12px;font-size:12px;margin:3px;display:inline-block;">{k}</span>'
        for k in notable_kws
    ) if notable_kws else "<span style='color:#999;font-size:12px;'>今週の新規候補なし</span>"

    # ソース別タブ
    def _section(title, icon, items):
        return f"""
        <h2 style="font-size:15px;color:#fff;background:#1E3A5F;padding:8px 16px;
                   border-radius:4px;margin:32px 0 12px;">{icon} {title}</h2>
        {_item_rows(items)}"""

    return f"""<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>技術トレンドレポート {report.generated_at[:10]}</title>
<style>
  body {{ font-family:"游ゴシック","Yu Gothic","Hiragino Kaku Gothic ProN",sans-serif;
         max-width:980px;margin:0 auto;padding:24px 16px 60px;background:#fff;color:#111; }}
  h1 {{ font-size:22px;color:#1E3A5F;border-bottom:3px solid #1E3A5F;padding-bottom:8px; }}
  a {{ text-decoration:none; }} a:hover {{ text-decoration:underline; }}
  .meta {{ font-size:12px;color:#6B7280;margin:4px 0 28px; }}
</style>
</head>
<body>

<h1>📡 技術トレンドレポート</h1>
<div class="meta">
  生成日時: {report.generated_at} ／
  GitHub {len(report.github_items)}件 ／
  Qiita {len(report.qiita_items)}件 ／
  Zenn {len(report.zenn_items)}件 ／
  npm {len(report.npm_items)}件
</div>

<h2 style="font-size:15px;color:#fff;background:#1E3A5F;padding:8px 16px;border-radius:4px;margin:0 0 12px;">
  📊 カテゴリ別件数
</h2>
<div style="background:#F9FAFB;border-radius:6px;padding:16px 20px;margin-bottom:8px;">
{cat_summary}
</div>

<h2 style="font-size:15px;color:#fff;background:#1E3A5F;padding:8px 16px;border-radius:4px;margin:32px 0 12px;">
  🔍 注目キーワード（config.py候補）
</h2>
<div style="padding:8px 0 4px;">{kw_chips}</div>

{_section("GitHub Trending", "🐙", report.github_items)}
{_section("Qiita 週間", "📝", report.qiita_items)}
{_section("Zenn トレンド", "✍️", report.zenn_items)}
{_section("npm 週間DL上位", "📦", report.npm_items)}

<div style="margin-top:48px;font-size:11px;color:#9CA3AF;text-align:center;">
  自動生成 by trend_collector — {report.generated_at}
</div>
</body></html>"""


# ─── Excel出力 ────────────────────────────────────────────────────────
def _save_excel(report: TrendReport, path: Path) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    CAT_COLORS = {
        "🤖 AI / LLM":          "D1FAE5",
        "📦 組み込み・エッジ":   "DBEAFE",
        "☁️ インフラ・クラウド": "FEF9C3",
        "🎨 フロントエンド":     "FCE7F3",
        "🔐 セキュリティ":       "FEE2E2",
        "⚙️ DevOps / MLOps":    "F3E8FF",
        "📊 データ・分析":       "FFEDD5",
        "🦀 言語・ランタイム":   "F0FDF4",
        "💡 その他":             "F9FAFB",
    }

    wb = openpyxl.Workbook()

    # ── シート1: 全アイテム ──
    ws = wb.active
    ws.title = "トレンド一覧"
    headers = ["ソース","カテゴリ","タイトル","URL","説明","言語/タグ","スター/いいね/DL","注目"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.font = Font(bold=True, color="FFFFFF", name="游ゴシック")
        c.fill = PatternFill("solid", fgColor="1E3A5F")
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20

    all_items = report.github_items + report.qiita_items + report.zenn_items + report.npm_items
    for ri, it in enumerate(all_items, 2):
        metric  = it.stars or it.likes or it.downloads
        tags    = "/".join(it.tags[:3]) if it.tags else it.language
        color   = CAT_COLORS.get(it.category, "F9FAFB")
        fill    = PatternFill("solid", fgColor=color)
        row_data = [it.source, it.category, it.title, it.url,
                    it.description[:80], tags, metric, "★" if it.is_notable else ""]
        for ci, v in enumerate(row_data, 1):
            cell = ws.cell(ri, ci, v)
            cell.fill = fill
            cell.font = Font(name="游ゴシック",
                             color="0563C1" if ci == 4 else "000000",
                             underline="single" if ci == 4 else None)
            if ci == 4 and str(v).startswith("http"):
                cell.hyperlink = str(v)
            cell.alignment = Alignment(vertical="center")
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 42
    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 25

    # ── シート2: カテゴリ別サマリー ──
    from collections import Counter
    ws2 = wb.create_sheet("カテゴリ別サマリー")
    ws2.cell(1, 1, "カテゴリ").font = Font(bold=True, color="FFFFFF", name="游ゴシック")
    ws2.cell(1, 1).fill = PatternFill("solid", fgColor="1E3A5F")
    ws2.cell(1, 2, "件数").font  = Font(bold=True, color="FFFFFF", name="游ゴシック")
    ws2.cell(1, 2).fill = PatternFill("solid", fgColor="1E3A5F")
    ws2.cell(1, 3, "上位タイトル").font = Font(bold=True, color="FFFFFF", name="游ゴシック")
    ws2.cell(1, 3).fill = PatternFill("solid", fgColor="1E3A5F")
    cat_items: dict[str, list] = {}
    for it in all_items:
        cat_items.setdefault(it.category, []).append(it)
    for ri, (cat, citems) in enumerate(sorted(cat_items.items(), key=lambda x: -len(x[1])), 2):
        top = " / ".join(i.title for i in sorted(citems, key=lambda x: -(x.stars or x.likes or x.downloads))[:3])
        ws2.cell(ri, 1, cat).font  = Font(name="游ゴシック")
        ws2.cell(ri, 2, len(citems)).font = Font(name="游ゴシック")
        ws2.cell(ri, 3, top).font  = Font(name="游ゴシック")
        color = CAT_COLORS.get(cat, "F9FAFB")
        for ci in range(1, 4):
            ws2.cell(ri, ci).fill = PatternFill("solid", fgColor=color)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 8
    ws2.column_dimensions["C"].width = 70

    # ── シート3: config更新候補 ──
    ws3 = wb.create_sheet("config更新候補")
    for ci, h in enumerate(["新規キーワード候補", "推定カテゴリ", "確認"], 1):
        ws3.cell(1, ci, h).font = Font(bold=True, color="FFFFFF", name="游ゴシック")
        ws3.cell(1, ci).fill   = PatternFill("solid", fgColor="92400E")
    notable = _extract_notable_keywords(all_items)
    for ri, kw in enumerate(notable, 2):
        cat = "ai_ml"
        kw_l = kw.lower()
        if any(x in kw_l for x in ["ros","robot","lidar"]): cat = "robot"
        elif any(x in kw_l for x in ["esp32","rtos","fpga","firmware","embed"]): cat = "embedded"
        elif any(x in kw_l for x in ["iot","matter"]): cat = "iot"
        elif any(x in kw_l for x in ["k8s","terraform","docker"]): cat = "infra"
        ws3.cell(ri, 1, kw).font  = Font(name="游ゴシック")
        ws3.cell(ri, 2, cat).font = Font(name="游ゴシック")
        ws3.cell(ri, 3, "").font  = Font(name="游ゴシック")
        for ci in range(1, 4):
            ws3.cell(ri, ci).fill = PatternFill("solid", fgColor="FEF3C7")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 20

    wb.save(path)
    log.info(f"  Excel保存 → {path}")


# ─── メイン ──────────────────────────────────────────────────────────
def run() -> TrendReport:
    log.info("=" * 55)
    log.info("技術トレンド収集開始（AI・エンジニアリング全般）")
    log.info("=" * 55)

    report = TrendReport(generated_at=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    report.github_items = collect_github_trending(since="weekly")
    time.sleep(1)
    report.qiita_items  = collect_qiita()
    time.sleep(1)
    report.zenn_items   = collect_zenn()
    time.sleep(1)
    report.npm_items    = collect_npm_trends()

    # 注目フラグを付与
    all_items = report.github_items + report.qiita_items + report.zenn_items + report.npm_items
    for it in all_items:
        it.is_notable = _has_notable(it.title + " " + " ".join(it.tags))
    report.new_keywords = _extract_notable_keywords(all_items)

    ts         = datetime.now().strftime("%Y%m%d")
    html_path  = OUTPUT_DIR / f"report_{ts}.html"
    excel_path = OUTPUT_DIR / f"report_{ts}.xlsx"
    html_path.write_text(_render_html(report), encoding="utf-8")
    log.info(f"  HTML保存 → {html_path}")
    _save_excel(report, excel_path)

    total = len(all_items)
    log.info(f"\n✅ 完了: {total} 件収集 / 注目KW {len(report.new_keywords)} 件")
    if report.new_keywords:
        log.info(f"🔍 注目: {', '.join(report.new_keywords)}")
    return report


if __name__ == "__main__":
    run()
